
"""
Excel generator for the Class Scheduling Program.
"""
import logging
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class ExcelGenerator:
    """Handles Excel report generation with styled tables."""

    def __init__(self, db_manager=None):
        self.db_manager = db_manager
        if not OPENPYXL_AVAILABLE:
            logging.error("openpyxl not available. Please install it with: pip install openpyxl")

    def generate_excel(self, title, headers, data, filename="report.xlsx"):
        """
        Generates an Excel file with a styled table from structured data.
        """
        if not OPENPYXL_AVAILABLE:
            return "openpyxl kütüphanesi bulunamadı. Lütfen 'pip install openpyxl' komutuyla yükleyin."

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = title[:31] # Sheet title limit is 31 chars

            # --- Styling ---
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # --- Title ---
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            title_cell = ws.cell(row=1, column=1, value=title)
            title_cell.font = Font(bold=True, size=18)
            title_cell.alignment = Alignment(horizontal="center")

            # --- Headers ---
            ws.append(headers)
            # Re-accessing cells to apply style after appending
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=2, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

            # --- Data ---
            for row_data in data:
                ws.append(row_data)

            # --- Post-process styling and formatting ---
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = center_align
            
            # Adjust column widths and row heights
            for col_idx in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 25
            for row_idx in range(2, ws.max_row + 1):
                ws.row_dimensions[row_idx].height = 40

            wb.save(filename)
            return f"Excel raporu başarıyla oluşturuldu: {filename}"

        except Exception as e:
            error_msg = f"Excel oluşturma hatası: {str(e)}"
            logging.error(error_msg)
            return error_msg
